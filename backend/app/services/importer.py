"""Turn an uploaded company file into a sending queue.

Three jobs, in order: read whatever columns the file happens to have, split
cells that hold more than one address, and collapse what is genuinely the same
address twice.

The guiding rule is that nothing is invented and nothing is silently dropped. A
blank website stays blank. A row that cannot be emailed is kept in the queue
marked `skipped` with the reason written on it, because a company that
disappears between the spreadsheet and the screen is a company you will spend
an afternoon looking for.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field
from functools import lru_cache

from .geo import tidy

log = logging.getLogger("bevigrow.import")

# Deliberately permissive. This decides "is this an address at all", not "will
# it accept mail" — only the mail server knows that, and a rule strict enough
# to argue with a real address is worse than one that lets it through.
EMAIL_RE = re.compile(r"^[^@\s,;]+@[^@\s,;]+\.[A-Za-z]{2,}$")

# One cell, several addresses: "info@x.com; sales@x.com" or comma/pipe/newline.
ADDRESS_SPLIT = re.compile(r"[;,|/\n\r\t]+| {2,}")

# Legal-entity suffixes, removed only when comparing two names.
_SUFFIXES = {
    "gmbh", "ltd", "limited", "llc", "inc", "incorporated", "bv", "b.v", "nv",
    "srl", "sarl", "sa", "ag", "as", "aps", "oy", "ab", "plc", "pty", "pte",
    "co", "corp", "corporation", "company", "kk", "kabushiki", "llp", "lp",
    "gk", "sl", "spa", "s.p.a", "kg", "ug", "ou", "sro", "doo", "dmcc", "fzco",
    "fze", "trading", "trade",
}

_PUNCT = re.compile(r"[^\w\s]")


# Memoised because the same handful of names is normalised over and over: the
# country note reads one key per send *and* one per log row, and a company
# appears in both. The work is a regex substitution and a case fold — trivial
# once, and most of a request when it is done six thousand times.
@lru_cache(maxsize=4096)
def normalize_company(raw: str | None) -> str:
    """Comparison key for a company name.

    Punctuation, case and legal suffixes are noise when deciding whether two
    rows are the same business: "ABC Coffee GmbH", "ABC Coffee, GmbH." and "abc
    coffee" all reduce to `abc coffee`. The original spelling is kept on the
    row — this is only ever the key.
    """
    text = _PUNCT.sub(" ", tidy(raw)).casefold()
    words = [w for w in text.split() if w and w not in _SUFFIXES]
    return " ".join(words) or text.strip()


def normalize_email(raw: str | None) -> str:
    return tidy(raw).casefold()


# What people write in a cell that has nothing in it. Left as-is, "N/A" becomes
# a website, gets stored, and one day somebody clicks it.
_BLANKS = {"n/a", "na", "n.a.", "-", "--", "—", "none", "null", "nil", "tbd", "?", "."}

# A URL pasted out of a document arrives as markdown: [www.x.de](https://www.x.de)
_MARKDOWN_LINK = re.compile(r"^\[([^\]]+)\]\((https?://[^)]+)\)$")


def clean_cell(raw: str | None) -> str:
    """One cell, as a person meant it.

    Placeholders for "I don't have this" become empty, because empty is what
    they mean and the template's fallbacks already handle empty properly.
    """
    value = tidy(raw)
    if not value:
        return ""
    if value.casefold() in _BLANKS:
        return ""
    link = _MARKDOWN_LINK.match(value)
    if link:
        return link.group(2)
    return value


def domain_of(email: str | None, website: str | None = None) -> str | None:
    """The domain to compare on: from the address first, the website second."""
    if email and "@" in email:
        return email.rsplit("@", 1)[-1].strip().casefold() or None
    if website:
        host = tidy(website).casefold()
        host = re.sub(r"^https?://", "", host)
        host = host.split("/")[0].strip()
        host = re.sub(r"^www\.", "", host)
        return host or None
    return None


# Mailboxes anyone can open. A domain is a good name for a company right up
# until the company uses gmail — and in this trade a great many do. Grouping
# on the domain would then file every unrelated firm with a gmail address as
# one business, so for these the company name is the better key.
FREE_MAIL = {
    "gmail.com", "googlemail.com", "yahoo.com", "yahoo.co.in", "yahoo.co.uk",
    "hotmail.com", "outlook.com", "live.com", "msn.com", "aol.com",
    "icloud.com", "me.com", "mac.com", "protonmail.com", "proton.me",
    "gmx.com", "gmx.de", "gmx.net", "web.de", "mail.com", "zoho.com",
    "yandex.com", "yandex.ru", "rediffmail.com", "sify.com", "vsnl.net",
    "bsnl.in", "airtelmail.in", "163.com", "126.com", "qq.com", "sina.com",
    "naver.com", "daum.net", "t-online.de", "orange.fr", "free.fr",
    "libero.it", "bol.com.br", "uol.com.br", "terra.com.br",
}


@lru_cache(maxsize=8192)
def company_key(
    email: str | None,
    website: str | None,
    name: str | None,
    location: str | None = None,
) -> str:
    """What counts as "the same company" for grouping.

    The mail domain when it identifies a business, the company name when it
    does not. Everything that groups companies — the importer, the country
    note, the duplicate report — has to use this one answer, or two screens
    will disagree about how many companies there are and both will be
    defensible.

    The location belongs in the name half of that, and has to be supplied
    when it is held in its own column. The send history keeps the name and
    the location apart; the outreach log writes them into one string. Reading
    the two without reconciling them gave "VBS_TESTING" and "VBS_TESTING,
    Salem" different keys, so a company recorded in both places counted twice
    — and doubling is the one arithmetic error nobody forgives in a number
    that reports how many firms they have reached.

    It is joined rather than discarded because two firms of one name in two
    cities are two firms, and dropping the town would merge them.
    """
    domain = domain_of(email, website)
    if domain and domain not in FREE_MAIL:
        return domain
    label = (name or "").strip()
    place = (location or "").strip()
    # Appended unless the name already ends with it, which is how the log
    # composes the two. Testing for the word anywhere in the name was wrong:
    # "Yercaud Estates" in Yercaud contains its town in the middle of its own
    # name, so the location was skipped on one side and appended on the other,
    # and the same firm ended up with two keys.
    if place and not label.casefold().rstrip(" ,").endswith(place.casefold()):
        label = f"{label}, {place}" if label else place
    return normalize_company(label) or (domain or (email or "").casefold() or "?")


# --------------------------------------------------------------- the columns


# Every spelling of a heading seen in the wild, mapped to the field it means.
# Matching is on a normalised header, so "Company Name", "company_name" and
# "COMPANY-NAME" all arrive here as "company name".
FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "company_name": (
        "company", "company name", "companyname", "name", "business",
        "business name", "organisation", "organization", "firm", "roaster",
        "importer", "buyer", "supplier", "account",
    ),
    "email": (
        "email", "e mail", "email address", "emails", "mail", "contact email",
        "email 1", "email 2", "email 3", "primary email", "secondary email",
        "alt email", "alternate email", "info email", "general email",
        "email id", "emailid", "email_id",
    ),
    "website": (
        "website", "web", "url", "site", "web address", "webpage", "homepage", "domain",
        "website url", "website address", "web site", "web url", "webiste", "site url",
        "company website", "business website", "org website", "organization website",
    ),
    "country": ("country", "nation", "market"),
    "location": ("location", "city", "address", "region", "state", "town", "province"),
    "contact_person": (
        "contact person", "contact", "person", "contact name", "owner",
        "manager", "director", "full name", "first name",
    ),
    "linkedin": ("linkedin", "linked in", "linkedin url", "linkedin profile"),
    "contact_form": ("contact form", "form", "contact page", "contact url", "enquiry form"),
    "phone": (
        "phone", "telephone", "tel", "mobile", "whatsapp", "phone number",
        "contact number", "cell",
    ),
    "category": ("category", "type", "segment", "industry", "business type", "specialty"),
}

_HEADER_CLEAN = re.compile(r"[^a-z0-9]+")


def _normalize_header(raw: str) -> str:
    return _HEADER_CLEAN.sub(" ", (raw or "").strip().casefold()).strip()


def map_headers(headers: list[str]) -> dict[int, str]:
    """Which column means what. Unrecognised columns are kept, not discarded.

    Returns {column index -> field name}. A column that matches nothing maps to
    `extra:<original heading>` and is stored as JSON on the row: the file's
    author put it there deliberately, and the template may well reference it.
    """
    lookup: dict[str, str] = {}
    for field_name, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            lookup[alias] = field_name

    mapping: dict[int, str] = {}
    for index, header in enumerate(headers):
        key = _normalize_header(header)
        if not key:
            continue
        if key in lookup:
            mapping[index] = lookup[key]
            continue
        # "email 2", "work email", "company website" — a heading that contains
        # a known one. Longest alias wins so "contact email" beats "contact".
        hit = max(
            (alias for alias in lookup if alias in key),
            key=len,
            default=None,
        )
        mapping[index] = lookup[hit] if hit else f"extra:{tidy(header)}"
    return mapping


# ------------------------------------------------------------------ the rows


@dataclass
class ParsedRow:
    """One address, ready to become a queue entry.

    One per mailbox, so each gets its own message. Several rows may share a
    company; the outreach log joins them when it writes them down.
    """

    position: int
    company_name: str
    email: str | None = None
    contact_person: str | None = None
    website: str | None = None
    country: str | None = None
    location: str | None = None
    linkedin: str | None = None
    contact_form: str | None = None
    phone: str | None = None
    category: str | None = None
    extra: dict[str, str] = field(default_factory=dict)
    skip_reason: str | None = None
    # What counts as "the same company" for grouping: the mail domain when
    # there is one, the normalised name otherwise. Set by the parser, because
    # it depends on the other rows around this one.
    group_key: str = ""

    @property
    def normalized_company(self) -> str:
        return self.group_key or normalize_company(self.company_name)

    @property
    def normalized_email(self) -> str:
        """The comparison key: the address, folded.

        Still written through the splitter rather than lowercased directly, so
        a cell that arrives with a stray separator or a display name around it
        reduces to the same key as the same address typed plainly.
        """
        return ",".join(_addresses_in(self.email or ""))

    @property
    def domain(self) -> str | None:
        return domain_of(self.email, self.website)


@dataclass
class ImportReport:
    """What the file turned into, in the words the summary will use."""

    rows: list[ParsedRow] = field(default_factory=list)
    file_rows: int = 0
    addresses: int = 0
    companies: int = 0
    multi_address_companies: int = 0
    duplicate_addresses: int = 0
    without_email: int = 0
    invalid_emails: int = 0
    possible_duplicates: list[str] = field(default_factory=list)
    unmapped_columns: list[str] = field(default_factory=list)
    # The same company name on more than one row. Not necessarily an error —
    # a firm with two offices is listed twice on purpose — but you should see
    # it before two people at the same company get near-identical letters.
    repeated_companies: list[str] = field(default_factory=list)
    # Different company names at one address. Usually a group with several
    # trading names, sometimes the same business entered twice. Reported so a
    # person decides, because only a person can.
    shared_locations: list[str] = field(default_factory=list)


def _rows_from_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet = book.active
        return [
            ["" if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        book.close()


def _rows_from_xls(data: bytes) -> list[list[str]]:
    """Excel 97-2003. Still what a lot of exports and old address books are."""
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    return [
        ["" if sheet.cell_value(r, c) is None else str(sheet.cell_value(r, c)).strip()
         for c in range(sheet.ncols)]
        for r in range(sheet.nrows)
    ]


def _rows_from_ods(data: bytes) -> list[list[str]]:
    """LibreOffice / OpenOffice, which is what a Google Sheet download can be."""
    from odf import table, text as odf_text
    from odf.opendocument import load

    doc = load(io.BytesIO(data))
    sheet = doc.spreadsheet.getElementsByType(table.Table)[0]
    rows: list[list[str]] = []
    for row in sheet.getElementsByType(table.TableRow):
        cells: list[str] = []
        for cell in row.getElementsByType(table.TableCell):
            value = "".join(
                str(p) for p in cell.getElementsByType(odf_text.P)
            ).strip()
            repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            # A run of identical cells is stored once with a repeat count;
            # expanding it keeps the columns lined up with the header.
            cells.extend([value] * min(repeat, 64))
        if any(c for c in cells):
            rows.append(cells)
    return rows


def _rows_from_text(data: bytes) -> list[list[str]]:
    """CSV, TSV, semicolon-separated, pipe-separated — whatever it turns out to be."""
    text: str | None = None
    # utf-8-sig first: Excel writes a byte-order mark that otherwise becomes
    # part of the first heading, and "﻿Company" matches no alias at all.
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Could not read that file as text.")

    lines = text.splitlines()
    first = lines[0] if lines else ""

    # The header line decides the delimiter, not csv.Sniffer.
    #
    # Sniffer gives up when rows have inconsistent field counts and raises —
    # and inconsistent counts are exactly what a German export produces, where
    # the file is semicolon-separated *and* an email cell holds
    # "info@x.de; sales@x.de". Falling back to a comma then reads the whole
    # line as one column and the import finds no addresses at all.
    #
    # A heading row almost never contains a delimiter that is not the
    # delimiter, so counting them there is both simpler and more reliable.
    counts = {d: first.count(d) for d in (",", ";", "\t", "|")}
    delimiter = max(counts, key=lambda d: counts[d])
    if counts[delimiter] == 0:
        try:
            delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","  # genuinely one column
    return [
        row
        for row in csv.reader(io.StringIO(text), delimiter=delimiter)
        if any(c.strip() for c in row)
    ]


def _read_table(data: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """Headers and rows from whatever was uploaded.

    Decided by what the bytes *are*, not what the name claims. A spreadsheet
    saved as "companies.csv" that is really an .xlsx is common enough — someone
    renamed it, or the browser did — and reading it as text produces one row of
    binary gibberish and a baffled user. The signature is unambiguous:

        PK..            a zip: .xlsx, .xlsm, .ods
        D0 CF 11 E0     OLE2:  .xls from Excel 97-2003
        anything else   text:  csv, tsv, txt, whatever the separator
    """
    name = (filename or "").lower()
    head = data[:8]

    def fail(kind: str, exc: Exception) -> ValueError:
        log.warning("Import failed for %s (%s): %s", filename, kind, exc)
        return ValueError(
            f"That file looked like {kind} but could not be read. "
            "If it opens in Excel, use File → Save As → CSV and upload that."
        )

    rows: list[list[str]] = []
    if head[:2] == b"PK":
        # Zip container: an .ods declares itself in the first entry's name.
        if b"opendocument.spreadsheet" in data[:4096] or name.endswith(".ods"):
            try:
                rows = _rows_from_ods(data)
            except ImportError as exc:
                raise ValueError(
                    "This build cannot read .ods files. Save the sheet as .xlsx or CSV."
                ) from exc
            except Exception as exc:  # noqa: BLE001
                raise fail("an OpenDocument sheet", exc) from exc
        else:
            try:
                rows = _rows_from_xlsx(data)
            except Exception as exc:  # noqa: BLE001
                raise fail("an Excel workbook", exc) from exc
    elif head[:4] == b"\xd0\xcf\x11\xe0":
        try:
            rows = _rows_from_xls(data)
        except ImportError as exc:
            raise ValueError(
                "This build cannot read the old .xls format. Open it in Excel and "
                "use File → Save As → .xlsx or CSV."
            ) from exc
        except Exception as exc:  # noqa: BLE001
            raise fail("an old Excel file", exc) from exc
    else:
        try:
            rows = _rows_from_text(data)
        except Exception as exc:  # noqa: BLE001
            raise fail("a text or CSV file", exc) from exc

    if not rows:
        return [], []
    # Trailing blank columns are normal in exported sheets; a header row of
    # ["Company", "Email", "", ""] would otherwise create two unnamed fields.
    header = [h.strip() for h in rows[0]]
    while header and not header[-1]:
        header.pop()
    return header, rows[1:]


def _addresses_in(cell: str) -> list[str]:
    """Every address in one cell, in order, without repeats.

    Reads the cell exactly as the sender will. It used to apply its own,
    narrower rule and threw away addresses the mail path would have handled
    perfectly: "info@hakanfoods.com - delivery delayed" and "Purvi LLC
    <purvillc@purvigroup.com>" were both discarded, so those companies were
    silently never written to at all — a worse outcome than a bounce, because
    a bounce at least tells you.

    Sharing the one cleaner also means the address that goes in the queue is
    character for character the address that goes in the envelope.
    """
    from .sender import recipients

    good, _rejected = recipients(cell or "")
    return good


def _join(unions: dict[str, str], domain: str, place: str | None) -> str:
    """Return the one group these two keys share, tying them together.

    A firm can be recognised two ways — by its mail domain, or by its name at
    its location — and the same firm may be recognised each way on different
    rows of the same file. Both keys are pointed at whichever group was seen
    first, so the second row joins the first rather than starting its own.
    """
    known = unions.get(domain) or (unions.get(place) if place else None)
    group = known or domain
    unions[domain] = group
    if place:
        unions[place] = group
    return group


def parse(data: bytes, filename: str) -> ImportReport:
    """Read the file into queue rows — one per address.

    A company with three mailboxes becomes three rows sharing a company key, so
    each gets its own email, its own result and its own line in the log, while
    the summary can still say "three addresses at one company".
    """
    headers, raw_rows = _read_table(data, filename)
    if not headers:
        raise ValueError("That file has no header row.")

    mapping = map_headers(headers)
    report = ImportReport()
    report.unmapped_columns = sorted(
        {v.split(":", 1)[1] for v in mapping.values() if v.startswith("extra:")}
    )

    seen_addresses: dict[str, str] = {}          # address -> the group that claimed it
    group_addresses: dict[str, set[str]] = {}    # group key -> its addresses
    row_by_group: dict[str, ParsedRow] = {}      # group key -> the one row it gets
    unions: dict[str, str] = {}                  # domain/name-place key -> its group
    names_by_norm: dict[str, set[str]] = {}      # normalised name -> spellings seen
    rows_by_norm: dict[str, int] = {}            # normalised name -> how many rows
    companies_at: dict[str, set[str]] = {}       # normalised address -> companies there
    groups_by_norm: dict[str, set[str]] = {}     # normalised name -> group keys
    position = 0

    for raw in raw_rows:
        values = {"extra": {}}
        emails: list[str] = []
        for index, cell in enumerate(raw):
            field_name = mapping.get(index)
            if not field_name:
                continue
            text = clean_cell(str(cell)) if cell is not None else ""
            if not text:
                continue
            if field_name == "email":
                emails.extend(_addresses_in(text))
            elif field_name.startswith("extra:"):
                values["extra"][field_name.split(":", 1)[1]] = text
            else:
                values[field_name] = text

        company = values.get("company_name") or ""
        # A location typed into the company-name cell is taken back out.
        #
        # Files often carry "Spice Star Foodstuff Trading LLC, Dubai" in the
        # name column *and* "Dubai" in the location column. The greeting reads
        # the name, so the letter would open "Dear Spice Star Foodstuff
        # Trading LLC, Dubai Team," — which no person would write and which
        # reads immediately as a mailshot.
        #
        # Only removed when the location column actually says the same thing:
        # a name genuinely ending in a place, with nothing to confirm it
        # against, is left exactly as the file has it. The log still shows the
        # name with the location after it — that is written when the row is
        # logged, from the two columns kept apart here.
        place = clean_cell(values.get("location"))
        if company and place:
            trimmed = company.rstrip()
            for suffix in (f", {place}", f" - {place}", f" ({place})", f", {place},"):
                if trimmed.casefold().endswith(suffix.casefold()):
                    company = trimmed[: -len(suffix)].rstrip(" ,-").strip() or trimmed
                    break
        if not company and not emails and not values.get("website"):
            continue  # a blank line in the sheet, not a company
        report.file_rows += 1
        if not company:
            # Filed under something findable rather than refused.
            company = values.get("website") or (emails[0] if emails else "Unnamed company")

        # A company, for grouping purposes, is its mail domain — or the same
        # name at the same place.
        #
        # The domain alone is not enough. "Bombay Foodstuff Trading Co. LLC,
        # Al Ras, Deira, Dubai" was listed twice with bombay@eim.ae and
        # info@bombayfoodstuff.com: one firm, two domains, and grouping on the
        # domain kept them apart. So an identical name at an identical
        # location in the same country groups too.
        #
        # The name is never enough on its own, which is the other half of the
        # rule: grouping on names alone once merged "ABC Coffee" in Japan with
        # "ABC Coffee GmbH" in Germany, two unrelated businesses. Requiring the
        # location and the country to match as well is what makes it safe —
        # two firms of the same name in the same city are a coincidence nobody
        # in this trade has hit, and a person can still split them by hand.
        norm = normalize_company(company)
        where = _PUNCT.sub(' ', tidy(values.get('location'))).casefold()
        where = ' '.join(where.split())
        place_key = None
        if norm and where:
            place_key = f"{norm}@@{where}@@{normalize_email(values.get('country'))}"
        group = company_key(
            emails[0] if emails else None,
            values.get("website"),
            company,
            values.get("location"),
        ) or norm
        # Two keys can name the same firm — the domain on one row, the
        # name-and-place on another. Whichever was seen first wins, so both
        # rows land in one group instead of two that never meet.
        group = _join(unions, group, place_key)
        names_by_norm.setdefault(norm, set()).add(company)
        rows_by_norm[norm] = rows_by_norm.get(norm, 0) + 1
        if where:
            companies_at.setdefault(where, set()).add(company)
        groups_by_norm.setdefault(norm, set()).add(group)
        group_addresses.setdefault(group, set())

        # Addresses that are not addresses: kept, marked, never sent to.
        raw_email_cells = [
            tidy(str(raw[i])) for i, f in mapping.items() if f == "email" and i < len(raw) and raw[i]
        ]
        junk = [c for c in raw_email_cells if c and not _addresses_in(c)]
        if junk:
            report.invalid_emails += 1

        # One place that builds a row, so the two branches below cannot drift.
        def build(email: str | None, skip_reason: str | None = None) -> ParsedRow:
            nonlocal position
            position += 1
            return ParsedRow(
                position=position,
                company_name=company,
                email=email,
                skip_reason=skip_reason,
                contact_person=values.get("contact_person"),
                website=values.get("website"),
                country=values.get("country"),
                location=values.get("location"),
                linkedin=values.get("linkedin"),
                contact_form=values.get("contact_form"),
                phone=values.get("phone"),
                category=values.get("category"),
                extra=dict(values["extra"]),
                group_key=group,
            )

        if not emails:
            report.rows.append(
                build(
                    None,
                    skip_reason=(
                        f"No usable email address in the file (found {junk[0]!r})"
                        if junk
                        else "No email address in the file"
                    ),
                )
            )
            report.without_email += 1
            continue

        fresh: list[str] = []
        for address in emails:
            folded = address.casefold()
            if folded in seen_addresses:
                # The same mailbox twice — in one row, or two rows apart. One
                # mailbox gets one message however many times it is listed.
                report.duplicate_addresses += 1
                continue
            seen_addresses[folded] = group
            group_addresses[group].add(folded)
            fresh.append(address)

        if not fresh:
            continue

        # One queue row per address, and so one email per address.
        #
        # These briefly shared an envelope: a firm listing info@ and sales@
        # received a single message addressed to both. It read better, and it
        # was wrong. Providers reject a whole request over one malformed
        # recipient, so a single bad address at a company stopped the good one
        # hearing from us at all — and a bounce then said nothing about which
        # of the two had failed. Separate messages fail separately.
        #
        # The outreach *log* still shows one row per company; that is done
        # when the row is written, not here. The queue is about envelopes.
        for address in fresh:
            report.rows.append(build(address))

    # Addresses, not rows: one row can now carry several.
    report.addresses = sum(len(_addresses_in(r.email)) for r in report.rows if r.email)
    report.companies = len(group_addresses)
    report.multi_address_companies = sum(1 for v in group_addresses.values() if len(v) > 1)

    # Names that read as the same business but sit on different domains.
    # Reported, never merged: "ABC Coffee" and "ABC Coffee GmbH" may be one
    # firm with two sites or two firms with similar names, and only a person
    # knows which. Guessing wrong either double-sends or drops a customer, so
    # both are queued and the pair is put in front of you instead.
    report.repeated_companies = sorted(
        f"{sorted(names_by_norm[norm])[0]} — {count} rows"
        for norm, count in rows_by_norm.items()
        if count > 1
    )
    report.shared_locations = sorted(
        f"{' / '.join(sorted(names))} — all at {address}"
        for address, names in companies_at.items()
        if len(names) > 1
    )

    for norm, groups in groups_by_norm.items():
        if len(groups) > 1:
            spellings = sorted(names_by_norm.get(norm, {norm}))
            report.possible_duplicates.append(
                f"{' / '.join(spellings)}  ({', '.join(sorted(groups))})"
            )

    return report
